# #############################################################################
#   
#   sunsynkGetMonthFigures.py
#
#   For a given month, get the daily figures for:
#   -   Generated (PV)
#   -   Load
#   -   Power Exported
#   -   Power Imported
#   -   Discharged Power from battery
#   -   Power to Charge the battery
#
#   These figures are saved to an Excel Spreadsheet in the current directory:
#       PV Monthly <YYYY-MM>.xlsx
# #############################################################################

import sunsynkAPI                               # library of routines to call Sunsynk API
from openpyxl import Workbook                   # openpyxl to create/edit/etc Excel files
from openpyxl.styles import Alignment, Font
from datetime import datetime
import click                                    # Prompt for parameters if not supplied


# not the proper way to do things, but set to True for debugging
debugFlag = False

# #############################################################################
#
#   Get the parameters.  Use the following to see them
#       python .\sunsynkGetMonthFigures.py --help
#
# #############################################################################
@click.group()
def init():
    # init function for click
    pass

# pylint: disable=no-value-for-parameter
@click.command()
@click.option('-u', '--username',
              prompt='Username:',
              help='Sunsynk username')
@click.option('-p', '--password',
              prompt='Password',
              help='Sunsynk password')
@click.option('-m', '--month',
              prompt='Month (YYYY-MM)',
              help='Enter month in YYYY-MM format')
def getMonthlyFigures(username, password, month):
    # #############################################################################
    #
    #   Call routines to get the monthly figures
    #
    # #############################################################################
    bearerToken = sunsynkAPI.getBearerToken(username, password)
    plantId = sunsynkAPI.getPlantId(bearerToken)
    monthlyStats = sunsynkAPI.getMonthlyStats (bearerToken, plantId, month)

    # #############################################################################
    #
    #   Create the Excel Workbook and set column headers
    #
    # #############################################################################
    wbOut = Workbook()
    row = 1
    wsOut = wbOut.active
    wsOut.cell(row=row, column=1, value = "Date").font = Font(bold=True)
    wsOut.column_dimensions['A'].width = 12
    wsOut.cell(row=row, column=2, value = "PV").font = Font(bold=True)
    wsOut.column_dimensions['B'].width = 11
    wsOut.cell(row=row, column=3, value = "Load").font = Font(bold=True)
    wsOut.column_dimensions['C'].width = 11
    wsOut.cell(row=row, column=4, value = "Export").font = Font(bold=True)
    wsOut.column_dimensions['D'].width = 11
    wsOut.cell(row=row, column=5, value = "Import").font = Font(bold=True)
    wsOut.column_dimensions['E'].width = 11
    wsOut.cell(row=row, column=6, value = "Discharge").font = Font(bold=True)
    wsOut.column_dimensions['F'].width = 11
    wsOut.cell(row=row, column=7, value = "Charge").font = Font(bold=True)
    wsOut.column_dimensions['G'].width = 11

    # #############################################################################
    #
    #   populate the spreadsheet with the figures
    #
    # #############################################################################
    row = 2
    for d in monthlyStats.keys():
        readingDate = datetime.strptime(d[0:10], '%Y-%m-%d')
        wsOut.cell(row=row, column=1, value = readingDate).number_format = format="dd-mmm-yy;@"
        wsOut.cell(row=row, column=2, value = float(monthlyStats[d]["PV"]))
        wsOut.cell(row=row, column=3, value = float(monthlyStats[d]["Load"]))
        wsOut.cell(row=row, column=4, value = float(monthlyStats[d]["Export"]))
        wsOut.cell(row=row, column=5, value = float(monthlyStats[d]["Import"]))
        wsOut.cell(row=row, column=6, value = float(monthlyStats[d]["Discharge"]))
        wsOut.cell(row=row, column=7, value = float(monthlyStats[d]["Charge"]))

        row += 1
    
    # #############################################################################
    #
    #   populate the spreadsheet with the figures
    #
    # #############################################################################
    saveFile = "PV Monthy " + month + ".xlsx"
    try:
        wbOut.save(saveFile)
        print ("Figures are in file " + saveFile)
    except:
        print("ERROR: Unable to save Excel file " + saveFile + " - file is open in Excel or otherwise locked")
    wbOut.close()
    return

# #############################################################################
if __name__ == "__main__":
    getMonthlyFigures()
    exit()
